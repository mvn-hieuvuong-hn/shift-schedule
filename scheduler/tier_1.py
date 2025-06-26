from ortools.sat.python import cp_model
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re
import calendar
from datetime import date
import os
import random

from members.tier_1_members import members
from weight.tier_1_weight_map import weight_map

scaled_weight_map = {
    'weekday': {k: int(v * 10) for k, v in weight_map['weekday'].items()},
    'weekend': {k: int(v * 10) for k, v in weight_map['weekend'].items()},
    'holiday': {k: int(v * 10) for k, v in weight_map['holiday'].items()}
}

# Hàm classify_day
def classify_day(d_obj, holiday_list_param):
    if d_obj in holiday_list_param:
        return 'holiday'
    elif d_obj.weekday() >= 5:
        return 'weekend'
    return 'weekday'

def generate_tier1_schedule_file(desired_month, desired_year, holiday_input_str, output_dir):
    
    try:
        # Tính toán các biến phụ thuộc vào members và weight_map
        num_members = len(members)
        sa_members_local = [m for m in members if "(SA)" in m]
        non_sa_members_local = [m for m in members if m not in sa_members_local]
        num_sa = len(sa_members_local)
        member_to_id = {m: i for i, m in enumerate(members)}
        sa_members_indices = [member_to_id[m] for m in sa_members_local]
        non_sa_members_indices = [member_to_id[m] for m in non_sa_members_local]
        all_total_night_shifts_vars = {}
        member_shift_counts = {}
        sa_weekday_ca4_counts = {} 

        # --- Xử lý ngày lễ ---
        holiday_list = []
        if holiday_input_str:
            day_strings = re.split(r'[, ]+', holiday_input_str)
            _, max_day_in_month = calendar.monthrange(desired_year, desired_month)
            for day_str in day_strings:
                try:
                    day_num = int(day_str)
                    if 1 <= day_num <= max_day_in_month:
                        holiday_list.append(date(desired_year, desired_month, day_num))
                    else:
                        print(f"Cảnh báo: Ngày '{day_str}' không hợp lệ cho tháng {desired_month}/{desired_year} và sẽ bị bỏ qua.")
                except ValueError:
                    print(f"Cảnh báo: '{day_str}' không phải là một số hợp lệ và sẽ bị bỏ qua.")
            holiday_list = sorted(list(set(holiday_list)))

        # --- Chuẩn bị dữ liệu lịch ---
        _, num_days = calendar.monthrange(desired_year, desired_month)
        date_list = [date(desired_year, desired_month, d) for d in range(1, num_days + 1)]

        weekdays_idx = [i for i, d in enumerate(date_list) if d.weekday() < 5 and d not in holiday_list]
        weekends_idx = [i for i, d in enumerate(date_list) if d.weekday() >= 5 or d in holiday_list]

        day_types = [classify_day(d, holiday_list) for d in date_list]
        monthly_total_weighted_hours_float = sum(
            sum(weight_map.get(day_types[i], {}).values()) for i in range(num_days)
        )

        total_shifts_per_type = {}
        for i, t in enumerate(day_types):
            for shift in weight_map.get(t, {}):
                total_shifts_per_type[(t, shift)] = total_shifts_per_type.get((t, shift), 0) + 1

        # --- Xây dựng mô hình CP-SAT ---
        model = cp_model.CpModel()
        shift_vars = {}
        all_shifts_by_day = {}

        # --- Khởi tạo ca ---
        ca123 = ["Ca 1", "Ca 2", "Ca 3"]
        shift_weekday_to_balance = ["Ca 1", "Ca 2", "Ca 3", "Ca 5", "Ca 6"]
        shift_weekend_holiday_to_balance = ["Ca 1", "Ca 2", "Ca 3", "Ca 4", "Ca 5", "Ca 6", "Ca 7", "Ca 8"]
        shift_day_weekend_holiday_to_balance = ["Ca 4", "Ca 5", "Ca 6", "Ca 7", "Ca 8"]

        for day_idx, day_type in enumerate(day_types):
            shifts = list(weight_map.get(day_type, {}).keys())
            all_shifts_by_day[day_idx] = shifts
            for shift in shifts:
                for m_idx in range(num_members):
                    shift_vars[(day_idx, shift, m_idx)] = model.NewBoolVar(f"shift_d{day_idx}_s{shift}_m{m_idx}")

        for day_idx, shifts in all_shifts_by_day.items():
            # --- Ràng buộc mỗi ca phải có 1 người ---
            for shift in shifts:
                model.AddExactlyOne(shift_vars[(day_idx, shift, m_idx)] for m_idx in range(num_members))
            # --- Ràng buộc mỗi người chỉ 1 ca/ngày ---
            for m_idx in range(num_members):
                model.AddAtMostOne(shift_vars[(day_idx, shift, m_idx)] for shift in shifts)

        # --- Ràng buộc không cho một người trực các ca trong tuần liền nhau ---
        for m_idx in range(num_members):
            # Chỉ xét các cặp ngày thường liên tiếp để giảm số ràng buộc
            for i in range(len(weekdays_idx) - 1):
                day_idx = weekdays_idx[i]
                next_day_idx = weekdays_idx[i + 1]
                # Đảm bảo hai ngày này liên tiếp nhau trong lịch
                if next_day_idx == day_idx + 1:
                    shifts_today = [s for s in shift_weekday_to_balance if s in all_shifts_by_day[day_idx]]
                    shifts_tomorrow = [s for s in shift_weekday_to_balance if s in all_shifts_by_day[next_day_idx]]
                    if shifts_today and shifts_tomorrow:
                        sum_shifts_d = sum(shift_vars[(day_idx, s, m_idx)] for s in shifts_today)
                        sum_shifts_d_plus_1 = sum(shift_vars[(next_day_idx, s, m_idx)] for s in shifts_tomorrow)
                        model.Add(sum_shifts_d + sum_shifts_d_plus_1 <= 1)

        # --- Ràng buộc không cho một người trực các ca 1,2,3 liền nhau ở các ngày trong tháng ---
        for m_idx in range(num_members):
            for day_idx in range(len(date_list) - 1):
                sum_shifts_today = sum(shift_vars[(day_idx, s, m_idx)] for s in ca123 if s in all_shifts_by_day[day_idx])
                sum_shifts_tomorrow = sum(shift_vars[(day_idx + 1, s, m_idx)] for s in ca123 if s in all_shifts_by_day[day_idx + 1])
                model.Add(sum_shifts_today + sum_shifts_tomorrow <= 1)

        # --- Ràng buộc: Trong 5 ngày liên tiếp, mỗi người tối đa 2 ca đêm Ca 1,2,3 ---
        window_size = 5
        max_night_shifts_in_window = 2
        for m_idx in range(num_members):
            for start in range(len(date_list) - window_size + 1):
                window_day_indices = range(start, start + window_size)
                night_shifts_in_window = [
                    shift_vars[(day_idx, shift, m_idx)]
                    for day_idx in window_day_indices
                    for shift in ca123
                    if shift in all_shifts_by_day[day_idx]
                ]
                if night_shifts_in_window:
                    model.Add(sum(night_shifts_in_window) <= max_night_shifts_in_window)

        # --- Ràng buộc nghỉ ngơi sau ca muộn/đêm (Ca 5, Ca 6 trong tuần; Ca 7, Ca 8 cuối tuần/lễ) 
        for day_idx in range(len(date_list) - 1):
            day_type = day_types[day_idx]
            if day_type == 'weekday':
                shifts_today_requiring_rest = [s for s in ("Ca 5", "Ca 6") if s in all_shifts_by_day[day_idx]]
            elif day_type in ('weekend', 'holiday'):
                shifts_today_requiring_rest = [s for s in ("Ca 7", "Ca 8") if s in all_shifts_by_day[day_idx]]
            else:
                shifts_today_requiring_rest = []
            if not shifts_today_requiring_rest:
                continue
            shifts_tomorrow = [s for s in ca123 if s in all_shifts_by_day[day_idx + 1]]
            if not shifts_tomorrow:
                continue
            for m_idx in range(num_members):
                for shift_today in shifts_today_requiring_rest:
                    shift_today_var = shift_vars[(day_idx, shift_today, m_idx)]
                    for next_day_shift in shifts_tomorrow:
                        shift_tomorrow_ca_var = shift_vars[(day_idx + 1, next_day_shift, m_idx)]
                        model.Add(shift_today_var + shift_tomorrow_ca_var <= 1)

        # --- Ràng buộc Ca 4 trong tuần chỉ SA và chia đều Ca 4 cho SA ---
        ca4_weekday_day_indices = [day_idx for day_idx in weekdays_idx if "Ca 4" in all_shifts_by_day[day_idx]]

        # Cấm non-SA trực Ca 4
        for day_idx in ca4_weekday_day_indices:
            for m_idx in non_sa_members_indices:
                model.Add(shift_vars[(day_idx, "Ca 4", m_idx)] == 0)

        # Chia đều Ca 4 cho SA
        total_ca4_weekday_shifts_count = len(ca4_weekday_day_indices)
        if total_ca4_weekday_shifts_count > 0 and num_sa > 0:
            base_ca4_per_sa = total_ca4_weekday_shifts_count // num_sa
            remainder_ca4 = total_ca4_weekday_shifts_count % num_sa
            for i, sa_idx in enumerate(sa_members_indices):
                ca4_shifts_for_this_sa = [
                    shift_vars[(day_idx, "Ca 4", sa_idx)]
                    for day_idx in ca4_weekday_day_indices
                ]
                sum_ca4_per_sa_var = model.NewIntVar(0, num_days, f"sa_{sa_idx}_ca4_weekday_sum")
                sa_weekday_ca4_counts[sa_idx] = sum_ca4_per_sa_var
                model.Add(sum_ca4_per_sa_var == sum(ca4_shifts_for_this_sa)) if ca4_shifts_for_this_sa else model.Add(sum_ca4_per_sa_var == 0)
                min_target = base_ca4_per_sa
                max_target = base_ca4_per_sa + (1 if remainder_ca4 > 0 else 0)
                model.Add(sum_ca4_per_sa_var >= min_target)
                model.Add(sum_ca4_per_sa_var <= max_target)

        # --- Ràng buộc cân bằng các ca trong tuần cho mỗi thành viên (Ca 1,2,3,5,6) ---
        total_occurrences_per_shift_weekday = {
            shift_type: sum(1 for day_idx in weekdays_idx if shift_type in all_shifts_by_day[day_idx])
            for shift_type in shift_weekday_to_balance
        }
        for m_idx in range(num_members):
            for shift_type in shift_weekday_to_balance:
                total_individual_shift_weekday_m = model.NewIntVar(0, num_days, f"total_{shift_type}_weekday_m{m_idx}")
                member_shift_counts[(m_idx, shift_type, 'weekday')] = total_individual_shift_weekday_m
                individual_shift_expr = [
                    shift_vars[(day_idx, shift_type, m_idx)]
                    for day_idx in weekdays_idx if shift_type in all_shifts_by_day[day_idx]
                ]
                model.Add(total_individual_shift_weekday_m == sum(individual_shift_expr)) if individual_shift_expr else model.Add(total_individual_shift_weekday_m == 0)
                total_occurrences = total_occurrences_per_shift_weekday[shift_type]
                avg_individual_shift_per_member = total_occurrences // num_members
                remainder_individual_shift = total_occurrences % num_members
                min_target = avg_individual_shift_per_member
                max_target = avg_individual_shift_per_member + (1 if remainder_individual_shift > 0 else 0)
                if shift_type in ("Ca 5", "Ca 6"):
                    max_target = avg_individual_shift_per_member + (1 if remainder_individual_shift > 0 else 0) + 1  
                model.Add(total_individual_shift_weekday_m >= min_target)
                model.Add(total_individual_shift_weekday_m <= max_target)
                # print(f" - Thành viên {members[m_idx]}: {shift_type} trong tuần [{min_target}-{max_target}]")

        # --- Ràng buộc cân bằng các ca cuối tuần/lễ cho mỗi thành viên (Ca 1,2,3,4,5,6,7,8) ---   
        total_occurrences_per_shift_weekend = {
            shift_type: sum(1 for day_idx in weekends_idx if shift_type in all_shifts_by_day[day_idx])
            for shift_type in shift_weekend_holiday_to_balance
        }
        for m_idx in range(num_members):
            for shift_type in shift_weekend_holiday_to_balance:
                total_individual_shift_weekend_m = model.NewIntVar(0, num_days, f"total_{shift_type}_weekend_m{m_idx}")
                member_shift_counts[(m_idx, shift_type, 'weekend/holiday')] = total_individual_shift_weekend_m
                individual_shift_expr = [
                    shift_vars[(day_idx, shift_type, m_idx)]
                    for day_idx in weekends_idx if shift_type in all_shifts_by_day[day_idx]
                ]
                model.Add(total_individual_shift_weekend_m == sum(individual_shift_expr)) if individual_shift_expr else model.Add(total_individual_shift_weekend_m == 0)
                total_occurrences = total_occurrences_per_shift_weekend[shift_type]
                avg_individual_shift_per_member = total_occurrences // num_members
                remainder_individual_shift = total_occurrences % num_members
                min_target = avg_individual_shift_per_member
                max_target = avg_individual_shift_per_member + (1 if remainder_individual_shift > 0 else 0)
                if shift_type in shift_day_weekend_holiday_to_balance:
                    max_target = avg_individual_shift_per_member + (1 if remainder_individual_shift > 0 else 0) + 1
                model.Add(total_individual_shift_weekend_m >= min_target)
                model.Add(total_individual_shift_weekend_m <= max_target)
                # print(f" - Thành viên {members[m_idx]}: {shift_type} cuối tuần [{min_target}-{max_target}]")      

        # --- Ràng buộc tổng ca 1,2,3 trong tuần cho mỗi thành viên ---
        total_possible_weekday_ca123_sum = sum(
            total_shifts_per_type.get(('weekday', shift_type), 0)
            for shift_type in ca123
        )
        avg_weekday_ca123_per_member_sum = total_possible_weekday_ca123_sum // num_members
        remainder_weekday_ca123_sum = total_possible_weekday_ca123_sum % num_members
        for m_idx in range(num_members):
            total_night_shifts_weekday_m = model.NewIntVar(0, num_days, f"total_night_weekday_m{m_idx}")
            all_total_night_shifts_vars[(m_idx, 'weekday')] = total_night_shifts_weekday_m
            model.Add(total_night_shifts_weekday_m == sum(
                shift_vars[(day_idx, shift, m_idx)]
                for day_idx in weekdays_idx
                for shift in ca123
                if shift in all_shifts_by_day[day_idx]
            ))
            min_target_night_weekday_sum = avg_weekday_ca123_per_member_sum
            max_target_night_weekday_sum = avg_weekday_ca123_per_member_sum + (1 if remainder_weekday_ca123_sum > 0 else 0)
            model.Add(total_night_shifts_weekday_m >= min_target_night_weekday_sum)
            model.Add(total_night_shifts_weekday_m <= max_target_night_weekday_sum)
            # print(f" - Thành viên {members[m_idx]}: Tổng ca 1,2,3 trong tuần [{min_target_night_weekday_sum}-{max_target_night_weekday_sum}]")

        # --- Ràng buộc tổng ca 1,2,3 cuối tuần & ngày lễ cho mỗi thành viên ---
        total_possible_weekend_holiday_ca123_sum = sum(
            total_shifts_per_type.get(('weekend', shift_type), 0) + total_shifts_per_type.get(('holiday', shift_type), 0)
            for shift_type in ca123
        )
        avg_weekend_holiday_ca123_per_member_sum = total_possible_weekend_holiday_ca123_sum // num_members
        remainder_weekend_holiday_ca123_sum = total_possible_weekend_holiday_ca123_sum % num_members
        for m_idx in range(num_members):
            total_night_shifts_weekend_m = model.NewIntVar(0, num_days, f"total_night_weekend_m{m_idx}")
            all_total_night_shifts_vars[(m_idx, 'weekend/holiday')] = total_night_shifts_weekend_m
            model.Add(total_night_shifts_weekend_m == sum(
                shift_vars[(day_idx, shift, m_idx)]
                for day_idx in weekends_idx
                for shift in ca123
                if shift in all_shifts_by_day[day_idx]
            ))
            min_target_night_weekend_sum = avg_weekend_holiday_ca123_per_member_sum
            max_target_night_weekend_sum = avg_weekend_holiday_ca123_per_member_sum + (1 if remainder_weekend_holiday_ca123_sum > 0 else 0)
            # print(f" - Thành viên {members[m_idx]}: Tổng ca 1,2,3 cuối tuần [{min_target_night_weekend_sum}-{max_target_night_weekend_sum}]")
            model.Add(total_night_shifts_weekend_m >= min_target_night_weekend_sum)
            model.Add(total_night_shifts_weekend_m <= max_target_night_weekend_sum)

        # --- Ràng buộc tổng số ca 1,2,3 trong tháng cho mỗi thành viên ---
        total_possible_overall_ca123_sum = sum(
            total_shifts_per_type.get(('weekday', shift_type), 0) +
            total_shifts_per_type.get(('weekend', shift_type), 0) +
            total_shifts_per_type.get(('holiday', shift_type), 0)
            for shift_type in ca123
        )
        avg_overall_ca123_per_member_sum = total_possible_overall_ca123_sum // num_members
        remainder_overall_ca123_sum = total_possible_overall_ca123_sum % num_members
        for m_idx in range(num_members):
            total_night_shifts_overall_m = model.NewIntVar(0, num_days, f"total_night_overall_m{m_idx}")
            model.Add(total_night_shifts_overall_m == sum(
                shift_vars[(day_idx, shift, m_idx)]
                for day_idx in range(len(date_list))
                for shift in ca123
                if shift in all_shifts_by_day[day_idx]
            ))
            min_target_night_overall_sum = avg_overall_ca123_per_member_sum
            max_target_night_overall_sum = avg_overall_ca123_per_member_sum + (1 if remainder_overall_ca123_sum > 0 else 0)
            # print(f" - Thành viên {members[m_idx]}: Tổng ca 1,2,3 [{min_target_night_overall_sum}-{max_target_night_overall_sum}]")    
            model.Add(total_night_shifts_overall_m >= min_target_night_overall_sum)
            model.Add(total_night_shifts_overall_m <= max_target_night_overall_sum)

        # --- Tính tổng giờ---
        total_hours = []
        for m_idx in range(num_members):
            total_hours_var = model.NewIntVar(0, 10000, f"total_hours_m{m_idx}")
            model.Add(total_hours_var == sum(
                shift_vars[(day_idx, shift, m_idx)] * scaled_weight_map.get(day_types[day_idx], {}).get(shift, 0)
                for day_idx in range(len(date_list))
                for shift in all_shifts_by_day[day_idx]
            ))
            total_hours.append(total_hours_var)

        # --- Ràng buộc và mục tiêu tổng giờ. SA hơn non-SA khoảng 20 giờ ---
        total_members_count = num_sa + len(non_sa_members_indices)
        x_float = (monthly_total_weighted_hours_float - num_sa * 20.0) / total_members_count
        target_non_sa_float = x_float
        target_sa_float = x_float + 20.0
        target_non_sa_scaled = int(target_non_sa_float * 10)
        target_sa_scaled = int(target_sa_float * 10)
        tolerance_scaled = 12 # Cho phép sai số +/- 1.2 giờ

        for m_idx in sa_members_indices:
            model.Add(total_hours[m_idx] >= target_sa_scaled - tolerance_scaled)
            model.Add(total_hours[m_idx] <= target_sa_scaled + tolerance_scaled)
        for m_idx in non_sa_members_indices:
            model.Add(total_hours[m_idx] >= target_non_sa_scaled - tolerance_scaled)
            model.Add(total_hours[m_idx] <= target_non_sa_scaled + tolerance_scaled)

        # --- Solver ---
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = 300
        status = solver.Solve(model)

        # --- Kết quả ---
        if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
            data = []
            holiday_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            weekend_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

            month_abbr = calendar.month_abbr 
            day_columns = [f"{d.day}-{month_abbr[d.month]}" for d in date_list]

            for m_idx in range(num_members):
                row_dict = {
                    "Member": members[m_idx],
                    "Total Hours": solver.Value(total_hours[m_idx]) / 10
                }
                for day_idx, col_name in enumerate(day_columns):
                    assigned_shift_for_day = next(
                        (s for s in all_shifts_by_day[day_idx] if solver.Value(shift_vars[(day_idx, s, m_idx)]) == 1),
                        ""
                    )
                    row_dict[col_name] = assigned_shift_for_day
                data.append(row_dict)

            df = pd.DataFrame(data)

            total_sum_hours = df["Total Hours"].sum()
            total_row_dict = {"Member": "Tổng cộng", "Total Hours": total_sum_hours}
            for col in df.columns:
                if col not in ["Member", "Total Hours"]:
                    total_row_dict[col] = ""
            df = pd.concat([df, pd.DataFrame([total_row_dict])], ignore_index=True)

            # Thay đổi logic lưu file để trả về cho Flask
            output_filename_base = f"Lich_truc_{desired_month}_{desired_year}.xlsx"
            output_filepath = os.path.join(output_dir, output_filename_base)

            workbook = Workbook()
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            sheet = workbook.create_sheet('LichTruc')

            for r in dataframe_to_rows(df, index=False, header=True):
                sheet.append(r)

            # Đặt độ rộng cột cho cột Member và Total Hours
            sheet.column_dimensions[get_column_letter(1)].width = 20
            sheet.column_dimensions[get_column_letter(2)].width = 10

            # Tạo sẵn danh sách màu cho từng ngày để tránh lặp lại phép kiểm tra
            day_fills = []
            for day_obj in date_list:
                if day_obj in holiday_list:
                    day_fills.append(holiday_fill)
                elif day_obj.weekday() >= 5:
                    day_fills.append(weekend_fill)
                else:
                    day_fills.append(None)

            # Đặt độ rộng và màu cho các cột ngày
            day_column_start_excel_idx = 3
            for col_offset, fill_color in enumerate(day_fills):
                col_letter = get_column_letter(day_column_start_excel_idx + col_offset)
                sheet.column_dimensions[col_letter].width = 8
                if fill_color:
                    for row_num in range(1, sheet.max_row + 1):
                        sheet[f'{col_letter}{row_num}'].fill = fill_color

            workbook.save(output_filepath)

            # Thống kê ca trực
            for m_idx in range(num_members):
                member_name = members[m_idx]
                weekday_123_count = solver.Value(all_total_night_shifts_vars[(m_idx, 'weekday')])
                weekend_123_count = solver.Value(all_total_night_shifts_vars[(m_idx, 'weekend/holiday')])
                print(f"- {members[m_idx]}: Tổng ca 1,2,3 trong tuần = {weekday_123_count}, cuối tuần = {weekend_123_count}")
                for shift_type in shift_weekday_to_balance:
                    weekday_count_var = member_shift_counts.get((m_idx, shift_type, 'weekday'))
                    weekday_count = solver.Value(weekday_count_var) if weekday_count_var is not None else 0
                    print(f"- {shift_type}: Trong tuần = {weekday_count}")
                if member_name in sa_members_local:
                    ca4_count_for_sa = solver.Value(sa_weekday_ca4_counts.get(m_idx, 0))
                    print(f"- Ca 4: Trong tuần (SA) = {ca4_count_for_sa}")
                for shift_type in shift_weekend_holiday_to_balance:
                    weekend_count_var = member_shift_counts.get((m_idx, shift_type, 'weekend/holiday'))
                    weekend_count = solver.Value(weekend_count_var) if weekend_count_var is not None else 0
                    print(f"- {shift_type}: Cuối tuần / Lễ = {weekend_count}")

            return True, "Lịch trực đã được tạo thành công!", output_filepath
        else:
            return False, f"Không tìm được lịch phù hợp. Trạng thái: {solver.StatusName(status)}", None
    except Exception as e:
        return False, f"Đã xảy ra lỗi trong quá trình tạo lịch: {e}", None
