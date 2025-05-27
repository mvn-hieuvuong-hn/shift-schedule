from ortools.sat.python import cp_model
import pandas as pd
from datetime import date, timedelta
import calendar
import re
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

# --- Trọng số ca ---
from tier_1_weight_map import weekday_shifts, weekend_shifts, holiday_shifts
from tier_1_members import members

weight_map = {
    'weekday': weekday_shifts,
    'weekend': weekend_shifts,
    'holiday': holiday_shifts
}

def get_month_year():
    while True:
        try:
            month = int(input("Nhập tháng (1-12): "))
            year = int(input("Nhập năm (ví dụ: 2025): "))
            if 1 <= month <= 12 and year > 0:
                return month, year
            else:
                print("Tháng phải từ 1 đến 12, năm phải là số dương.")
        except ValueError:
            print("Vui lòng nhập số nguyên.")

desired_month, desired_year = get_month_year()

def get_holiday_list(year, month):
    """
    Hỏi người dùng nhập các ngày lễ trong tháng.
    Nhập các ngày cách nhau bởi dấu phẩy hoặc khoảng trắng (ví dụ: 5, 10, 25).
    """
    holidays = []
    print(f"\nNhập các ngày lễ trong tháng {month}/{year} (ví dụ: 5, 10, 25), để trống nếu không có:")
    holiday_input_str = input("Các ngày lễ: ").strip()

    if not holiday_input_str:
        return [] # Trả về danh sách rỗng nếu không nhập gì
    
    day_strings = re.split(r'[, ]+', holiday_input_str)

    _, max_day_in_month = calendar.monthrange(year, month)

    for day_str in day_strings:
        try:
            day_num = int(day_str)
            if 1 <= day_num <= max_day_in_month:
                holidays.append(date(year, month, day_num))
            else:
                print(f"  Cảnh báo: Ngày '{day_str}' không hợp lệ cho tháng {month}/{year} và sẽ bị bỏ qua.")
        except ValueError:
            print(f"  Cảnh báo: '{day_str}' không phải là một số hợp lệ và sẽ bị bỏ qua.")
            
    # Loại bỏ các ngày trùng lặp và sắp xếp lại
    return sorted(list(set(holidays)))

# --- Cấu hình ngày lễ ---
holiday_list = get_holiday_list(desired_year, desired_month)

# --- Lấy danh sách ngày trong tháng ---
_, num_days = calendar.monthrange(desired_year, desired_month)
date_list = [date(desired_year, desired_month, d) for d in range(1, num_days + 1)]

def classify_day(d):
    if d in holiday_list:
        return 'holiday'
    elif d.weekday() >= 5:
        return 'weekend'
    return 'weekday'


monthly_total_weighted_hours_float = 0
for day_num in range(1, num_days + 1):
    current_date = date(desired_year, desired_month, day_num)
    day_type = classify_day(current_date)
    monthly_total_weighted_hours_float += sum(weight_map[day_type].values())


# monthly_total_weighted_hours_scaled = int(monthly_total_weighted_hours_float * 10)


# --- Thành viên ---
sa_members = [m for m in members if "(SA)" in m]
non_sa_members = [m for m in members if m not in sa_members]


# --- Tạo model ---
model = cp_model.CpModel()
num_members = len(members)
num_sa = len(sa_members)
num_non_sa = len(non_sa_members)
member_to_id = {m: i for i, m in enumerate(members)}

# --- Tạo biến: shift_vars[(day, shift, member)] = 0/1 ---
shift_vars = {}
all_shifts_by_day = {}

for day_idx, day in enumerate(date_list):
    day_type = classify_day(day)
    shifts = list(weight_map[day_type].keys())
    all_shifts_by_day[day_idx] = shifts
    for shift in shifts:
        for m in range(num_members):
            shift_vars[(day_idx, shift, m)] = model.NewBoolVar(f"shift_d{day_idx}_s{shift}_m{m}")

# --- Ràng buộc mỗi ca phải có 1 người ---
for day_idx in range(len(date_list)):
    shifts = all_shifts_by_day[day_idx]
    for shift in shifts:
        model.AddExactlyOne(shift_vars[(day_idx, shift, m)] for m in range(num_members))

# --- Ràng buộc mỗi người chỉ 1 ca/ngày ---
for day_idx in range(len(date_list)):
    shifts = all_shifts_by_day[day_idx]
    for m in range(num_members):
        model.AddAtMostOne(shift_vars[(day_idx, shift, m)] for shift in shifts)

# --- Ràng buộc Ca 4 trong tuần chỉ SA ---
for day_idx, day in enumerate(date_list):
    day_type = classify_day(day)
    if day_type == 'weekday':
        if "Ca 4" in all_shifts_by_day[day_idx]:
            for m in range(num_members):
                if members[m] not in sa_members:
                    model.Add(shift_vars[(day_idx, "Ca 4", m)] == 0)

# Lấy danh sách index các ngày trong tuần và cuối tuần hoặc lễ
weekdays_idx = [i for i, d in enumerate(date_list) if d.weekday() < 5 and d not in holiday_list]
weekends_idx = [i for i, d in enumerate(date_list) if d.weekday() >= 5 or d in holiday_list]

# --- Ràng buộc không cho người cùng lúc 3 ca 1,2,3 liền nhau ---
ca123 = ["Ca 1", "Ca 2", "Ca 3"]
for m in range(num_members):
    for day_idx in range(len(date_list) - 1):
        sum_shifts_d = sum(shift_vars[(day_idx, s, m)] for s in ca123 if s in all_shifts_by_day[day_idx])
        sum_shifts_d_plus_1 = sum(shift_vars[(day_idx + 1, s, m)] for s in ca123 if s in all_shifts_by_day[day_idx+1])
        model.Add(sum_shifts_d + sum_shifts_d_plus_1 <= 1)

# --- Ràng buộc nghỉ ngơi sau ca muộn/đêm (Ca 5, Ca 6 trong tuần; Ca 7, Ca 8 cuối tuần/lễ) ---
for m_idx in range(num_members):
    for day_idx in range(len(date_list) - 1): # Lặp qua các ngày trừ ngày cuối cùng
        current_day = date_list[day_idx]
        day_type = classify_day(current_day) # Phân loại ngày hôm nay

        # Các ca làm việc muộn/đêm của ngày hôm nay mà yêu cầu nghỉ ngơi ngày mai
        shifts_today_requiring_rest = []

        if day_type == 'weekday':
            # Nếu hôm nay là ngày trong tuần, xét Ca 5 và Ca 6
            if "Ca 5" in all_shifts_by_day[day_idx]:
                shifts_today_requiring_rest.append("Ca 5")
            if "Ca 6" in all_shifts_by_day[day_idx]:
                shifts_today_requiring_rest.append("Ca 6")
        elif day_type in ['weekend', 'holiday']: # Cuối tuần hoặc ngày lễ
            # Nếu hôm nay là cuối tuần hoặc ngày lễ, xét Ca 7 và Ca 8
            if "Ca 7" in all_shifts_by_day[day_idx]:
                shifts_today_requiring_rest.append("Ca 7")
            if "Ca 8" in all_shifts_by_day[day_idx]:
                shifts_today_requiring_rest.append("Ca 8")

        # Áp dụng ràng buộc cho các ca được xác định
        for shift_today in shifts_today_requiring_rest:
            shift_today_var = shift_vars[(day_idx, shift_today, m_idx)]
            
            # Thì hôm sau (ngày day_idx + 1) không được trực Ca 1, 2, 3
            for next_day_shift in ca123:
                if next_day_shift in all_shifts_by_day[day_idx + 1]: # Đảm bảo ca cấm có tồn tại vào ngày mai
                    shift_tomorrow_ca_var = shift_vars[(day_idx + 1, next_day_shift, m_idx)]
                    
                    # Ràng buộc: KHÔNG (trực ca muộn/đêm hôm nay VÀ trực Ca 1,2,3 ngày mai)
                    # Biểu diễn: shift_today_var + shift_tomorrow_ca_var <= 1
                    model.Add(shift_today_var + shift_tomorrow_ca_var <= 1)

# --- Ràng buộc tổng số ca 1, 2, 3 trong tuần cho mỗi thành viên ---
# Tính tổng số ca 1,2,3 có thể có vào các ngày trong tuần trong tháng đó
total_possible_weekday_ca123 = 0
for day_idx in weekdays_idx:
    for shift in ca123:
        if shift in all_shifts_by_day[day_idx]:
            total_possible_weekday_ca123 += 1

# Tính số ca 1,2,3 trung bình mỗi người nên có trong tuần
avg_weekday_ca123_per_member = total_possible_weekday_ca123 // num_members
remainder_weekday_ca123 = total_possible_weekday_ca123 % num_members

weekday_ca123_tolerance = 1 # Cho phép sai lệch +/- 1 ca

for m in range(num_members):
    # Tạo biến để lưu tổng số ca 1,2,3 trong tuần của thành viên m
    total_ca123_weekday_m = model.NewIntVar(0, num_days, f"total_ca123_weekday_m{m}")

    # Tính tổng các ca 1,2,3 trong tuần cho thành viên m
    weekday_night_shifts_expr = []
    for day_idx in weekdays_idx: # Chỉ xét các ngày trong tuần
        for shift in ca123:
            # Kiểm tra xem ca có tồn tại trong all_shifts_by_day[day_idx] không
            if shift in all_shifts_by_day[day_idx]:
                weekday_night_shifts_expr.append(shift_vars[(day_idx, shift, m)])

    model.Add(total_ca123_weekday_m == sum(weekday_night_shifts_expr))
    min_target_weekday_ca123 = avg_weekday_ca123_per_member
    max_target_weekday_ca123 = avg_weekday_ca123_per_member + (1 if m_idx < remainder_weekday_ca123 else 0)
    model.Add(total_ca123_weekday_m >= max(0, min_target_weekday_ca123 - weekday_ca123_tolerance))
    model.Add(total_ca123_weekday_m <= max_target_weekday_ca123 + weekday_ca123_tolerance)


individual_shift_tolerance = 1 # Cho phép sai lệch +/- 1 ca cho mỗi loại ca đêm


for m_idx in range(num_members):
    for shift_type in ca123:
        # Tính tổng số lần thành viên m_idx trực 'shift_type' trong các ngày trong tuần
        total_individual_shift_weekday_m = model.NewIntVar(0, num_days, f"total_{shift_type}_weekday_m{m_idx}")

        individual_shift_expr = []
        for day_idx in weekdays_idx: # Chỉ xét các ngày trong tuần
            if shift_type in all_shifts_by_day[day_idx]:
                individual_shift_expr.append(shift_vars[(day_idx, shift_type, m_idx)])

        model.Add(total_individual_shift_weekday_m == sum(individual_shift_expr))

        # Tính tổng số lần 'shift_type' này xuất hiện trong tất cả các ngày trong tuần
        total_occurrences_of_this_shift_weekday = 0
        for day_idx in weekdays_idx:
            if shift_type in all_shifts_by_day[day_idx]:
                total_occurrences_of_this_shift_weekday += 1
        
        # Nếu có ca này và có thành viên để phân bổ
        if total_occurrences_of_this_shift_weekday > 0 and num_members > 0:
            # Tính toán số lần trung bình mỗi thành viên nên trực ca này
            avg_individual_shift_per_member = total_occurrences_of_this_shift_weekday // num_members
            remainder_individual_shift = total_occurrences_of_this_shift_weekday % num_members

            # Thiết lập khoảng mục tiêu cho ca này
            min_target_individual_shift = avg_individual_shift_per_member
            max_target_individual_shift = avg_individual_shift_per_member + (1 if m_idx < remainder_individual_shift else 0)

            # Áp dụng ràng buộc với tolerance
            model.Add(total_individual_shift_weekday_m >= max(0, min_target_individual_shift - individual_shift_tolerance))
            model.Add(total_individual_shift_weekday_m <= max_target_individual_shift + individual_shift_tolerance)
            
# --- Ràng buộc tổng số ca 1,2,3 cuối tuần & ngày lễ cho mỗi thành viên ---
total_possible_weekend_holiday_ca123 = 0
for day_idx in weekends_idx:
    for shift in ca123:
        if shift in all_shifts_by_day[day_idx]:
            total_possible_weekend_holiday_ca123 += 1

avg_weekend_holiday_ca123_per_member = total_possible_weekend_holiday_ca123 // num_members
remainder_weekend_holiday_ca123 = total_possible_weekend_holiday_ca123 % num_members
weekend_holiday_ca123_tolerance = 1 # Cho phép sai lệch +/- 1 ca

for m in range(num_members):
    # Tạo biến để lưu tổng số ca 1,2,3 cuối tuần của thành viên m
    total_ca123_weekend_m = model.NewIntVar(0, num_days, f"total_ca123_weekend_m{m}")

    # Tính tổng các ca 1,2,3 cuối tuần cho thành viên m
    weekend_night_shifts_expr = []
    for day_idx in weekends_idx: # Chỉ xét các ngày cuối tuần và ngày lễ
        for shift in ca123:
            # Kiểm tra xem ca có tồn tại trong all_shifts_by_day[day_idx] không
            if shift in all_shifts_by_day[day_idx]:
                weekend_night_shifts_expr.append(shift_vars[(day_idx, shift, m)])

    min_target_weekend_holiday_ca123 = avg_weekend_holiday_ca123_per_member
    max_target_weekend_holiday_ca123 = avg_weekend_holiday_ca123_per_member + (1 if m_idx < remainder_weekend_holiday_ca123 else 0)
    model.Add(total_ca123_weekend_m >= max(0, min_target_weekend_holiday_ca123 - weekend_holiday_ca123_tolerance))
    model.Add(total_ca123_weekend_m <= max_target_weekend_holiday_ca123 + weekend_holiday_ca123_tolerance)

# --- Ràng buộc: Chia đều Ca 4 trong tuần cho 3 thành viên SA ---
sa_members_indices = [member_to_id[m] for m in sa_members]
total_ca4_weekday_shifts_count = 0
for day_idx in weekdays_idx:
    if "Ca 4" in all_shifts_by_day[day_idx]:
        total_ca4_weekday_shifts_count += 1

if total_ca4_weekday_shifts_count > 0:

    # Tính toán số ca 4 mục tiêu cho mỗi SA
    base_ca4_per_sa = total_ca4_weekday_shifts_count // num_sa
    remainder_ca4 = total_ca4_weekday_shifts_count % num_sa

    ca4_sa_tolerance = 1 # Cho phép sai lệch 1 ca

    for i, sa_idx in enumerate(sa_members_indices):
        # Tạo một biến để lưu tổng số Ca 4 của mỗi SA
        # sum_ca4_per_sa_var là một IntVar đại diện cho tổng số Ca 4 của SA này
        sum_ca4_per_sa_var = model.NewIntVar(0, num_days, f"sa_{sa_idx}_ca4_weekday_sum")

        ca4_shifts_for_this_sa = []
        for day_idx, day in enumerate(date_list):
            day_type = classify_day(day)
            if day_type == 'weekday' and "Ca 4" in all_shifts_by_day[day_idx]:
                ca4_shifts_for_this_sa.append(shift_vars[(day_idx, "Ca 4", sa_idx)])

        if ca4_shifts_for_this_sa:
            model.Add(sum_ca4_per_sa_var == sum(ca4_shifts_for_this_sa))
        else:
            model.Add(sum_ca4_per_sa_var == 0)
        min_target = base_ca4_per_sa
        max_target = base_ca4_per_sa + 1 if i < remainder_ca4 else base_ca4_per_sa

        model.Add(sum_ca4_per_sa_var >= max(0, min_target - ca4_sa_tolerance))
        model.Add(sum_ca4_per_sa_var <= max_target + ca4_sa_tolerance)


# --- Tính tổng giờ trực mỗi người ---
total_hours = []
for m in range(num_members):
    hours_expr = []
    for day_idx, day in enumerate(date_list):
        day_type = classify_day(day)
        for shift in all_shifts_by_day[day_idx]:
            weight = int(weight_map[day_type][shift] * 10)  # scale để dùng integer
            hours_expr.append(shift_vars[(day_idx, shift, m)] * weight)
    total_hours.append(model.NewIntVar(0, 1000, f"total_hours_m{m}"))
    model.Add(total_hours[m] == sum(hours_expr))

# --- Ràng buộc SA hơn non-SA khoảng 20 giờ ---
sa_ids = [member_to_id[m] for m in sa_members]
non_sa_ids = [member_to_id[m] for m in non_sa_members]

avg_sa = model.NewIntVar(0, 10000, "avg_sa")
avg_non_sa = model.NewIntVar(0, 10000, "avg_non_sa")

model.Add(avg_sa * len(sa_ids) == sum(total_hours[m] for m in sa_ids))
model.Add(avg_non_sa * len(non_sa_ids) == sum(total_hours[m] for m in non_sa_ids))

# SA hơn non-SA khoảng 20 giờ * 10 (scale)
model.Add(avg_sa >= avg_non_sa + 200)


# --- PHÂN BỔ GIỜ MỤC TIÊU --- #

# Tính giờ mục tiêu cho SA và non-SA (SA hơn ~20 giờ)

x_float = (monthly_total_weighted_hours_float - num_sa * 20.0) / (num_sa + num_non_sa)

target_non_sa_float = x_float
target_sa_float = x_float + 20.0

target_non_sa_scaled = int(target_non_sa_float * 10)
target_sa_scaled = int(target_sa_float * 10)

# Cho phép sai lệch 1 giờ (1 hour * 10 = 10 units)
tolerance_scaled = 10

print(f"\n🎯 Tổng giờ trực trong tháng: {monthly_total_weighted_hours_float:.1f} giờ")
print(f"🎯 Giờ mục tiêu cho mỗi SA: {target_sa_float:.1f} giờ")
print(f"🎯 Giờ mục tiêu cho mỗi non-SA: {target_non_sa_float:.1f} giờ")

# Áp ràng buộc cho từng thành viên
for m_idx in sa_ids:
    model.Add(total_hours[m_idx] >= target_sa_scaled - tolerance_scaled)
    model.Add(total_hours[m_idx] <= target_sa_scaled + tolerance_scaled)

for m_idx in non_sa_ids:
    model.Add(total_hours[m_idx] >= target_non_sa_scaled - tolerance_scaled)
    model.Add(total_hours[m_idx] <= target_non_sa_scaled + tolerance_scaled)

# --- Solver ---
solver = cp_model.CpSolver()
solver.parameters.max_time_in_seconds = 300
status = solver.Solve(model)

# --- Kết quả ---
if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
    data = []
    # Định nghĩa màu sắc
    holiday_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    weekend_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
    
    # Fill data and create the DataFrame
    for m_idx in range(num_members):
        row_dict = {}
        row_dict["Member"] = members[m_idx]
        row_dict["Total Hours"] = solver.Value(total_hours[m_idx]) / 10
        
        for day_idx, day in enumerate(date_list):
            assigned_shifts = [
                s for s in all_shifts_by_day[day_idx]
                if solver.Value(shift_vars[(day_idx, s, m_idx)]) == 1
            ]
            month_abbreviation = calendar.month_abbr[day.month]
            row_dict[f"{day.day}-{month_abbreviation}"] = assigned_shifts[0] if assigned_shifts else ""
        
        data.append(row_dict)
    
    df = pd.DataFrame(data)

    # --- SỬA ĐỔI PHẦN NÀY ---
    output_directory = "ShiftSchedule_Export"
    # Tạo thư mục nếu nó chưa tồn tại
    os.makedirs(output_directory, exist_ok=True) # exist_ok=True ngăn lỗi nếu thư mục đã có

    output_filename = os.path.join(output_directory, f"Lich_truc_{desired_month}_{desired_year}.xlsx")
    
    
    # Tạo một Workbook openpyxl trước
    workbook = Workbook()
    # Xóa sheet mặc định 'Sheet' nếu bạn không cần nó
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
    
    # Tạo sheet mới với tên mong muốn
    sheet_name = 'LichTruc'
    sheet = workbook.create_sheet(sheet_name) # Tạo sheet mới và lấy reference trực tiếp

    # Ghi DataFrame vào sheet đó
    # openpyxl.utils.dataframe.dataframe_to_rows là cách tốt để ghi vào sheet có sẵn
    for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True)):
        sheet.append(r)

    # Xác định chỉ số cột của các ngày
    day_column_start_excel_idx = 3 # Member (1), Total Hours (2). Ngày bắt đầu từ cột 3 (C)

    for col_offset, day_obj in enumerate(date_list):
        current_col_excel_idx = day_column_start_excel_idx + col_offset
        col_letter = get_column_letter(current_col_excel_idx)
        
        fill_color = None
        if day_obj in holiday_list:
            fill_color = holiday_fill
        elif day_obj.weekday() >= 5: # Saturday or Sunday
            fill_color = weekend_fill
        
        if fill_color:
            for row_num in range(1, sheet.max_row + 1):
                sheet[f'{col_letter}{row_num}'].fill = fill_color
    
    # Lưu workbook trực tiếp
    workbook.save(output_filename) # Lưu workbook trực tiếp
    
    print("\n✅ Lịch trực đã được tạo:")
    print(df) # Vẫn in DataFrame ra console
    print(f"📁 Đã lưu lịch trực vào file: {output_filename}")
else:
    print("Không tìm được lịch phù hợp.")
